import os
import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl

def get_xpath_to_excel(url, output_folder, output_filename):
    # ChromeDriverのセットアップ
    options = Options()
    options.add_argument("--headless")  # ヘッドレスモードでChromeを実行する場合はコメントアウトを解除
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    # ページを開く
    driver.get(url)

    # ページが完全に読み込まれるまで待機（最大で10秒）
    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_all_elements_located((By.XPATH, "//*")))

    # Excelファイルの作成
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "テキスト要素"
    sheet["B1"] = "要素のXPATH"
    sheet["C1"] = "HTML要素"
    sheet["D1"] = "ID要素"
    sheet["E1"] = "name要素"

    # すべての要素のXPathとテキストを取得してExcelに書き込む
    elements = driver.find_elements(By.XPATH, "//*")
    row = 2  # Excelの行番号
    for element in elements:
        text = element.text
        xpath = driver.execute_script(
            'function getPathTo(element) {'
            '    if (element.tagName === "HTML")'
            '        return "/HTML";'
            '    if (element === document.body)'
            '        return "/HTML/BODY";'
            ''
            '    var ix = 0;'
            '    var siblings = element.parentNode.childNodes;'
            '    for (var i = 0; i < siblings.length; i++) {'
            '        var sibling = siblings[i];'
            ''
            '        if (sibling === element)'
            '            return getPathTo(element.parentNode) + "/" + element.tagName + "[" + (ix + 1) + "]";'
            ''
            '        if (sibling.nodeType === 1 && sibling.tagName === element.tagName)'
            '            ix++;'
            '    }'
            '}'
            ''
            'return getPathTo(arguments[0]);',
            element
        )
        tag_name = element.tag_name
        element_id = element.get_attribute("id")
        element_name = element.get_attribute("name")
        sheet.cell(row=row, column=1, value=text)
        sheet.cell(row=row, column=2, value=xpath)
        sheet.cell(row=row, column=3, value=tag_name)
        sheet.cell(row=row, column=4, value=element_id)
        sheet.cell(row=row, column=5, value=element_name)
        row += 1

    # 出力先フォルダが存在しない場合は作成する
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 出力ファイルのパス
    output_path = os.path.join(output_folder, output_filename)

    # Excelファイルを保存
    workbook.save(output_path)

    # ブラウザを終了する
    driver.quit()


# 関数の呼び出し例
url = "https://qiita.com/drafts/new"
output_folder = r'C:\myworkspace\python\browser_automation'
current_datetime = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
output_filename = f"output_{current_datetime}.xlsx"
get_xpath_to_excel(url, output_folder, output_filename)
