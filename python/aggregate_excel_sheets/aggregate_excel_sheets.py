import os
from openpyxl import Workbook, load_workbook
import datetime


# folder_path：コピー対象のExcelがあるフォルダパス
# copy_sheet_name：コピー対象のシート名
# output_folder_path：集約ファイルの出力先
# output_file_name：集約ファイルのファイル名
def aggregate_excel_sheets(input_folder_path,input_sheet_name,output_folder_path,output_file_name):
    output_book = Workbook()

    for input_file_name in os.listdir(input_folder_path):
        if input_file_name.endswith('.xlsx'):
            
            input_file_path = os.path.join(input_folder_path, input_file_name)
            input_workbook = load_workbook(input_file_path)
            input_file_name_noext = os.path.splitext(os.path.basename(input_file_path))[0]
            
            if input_sheet_name in input_workbook.sheetnames:
                input_sheet = input_workbook[input_sheet_name]
                output_sheet = output_book.create_sheet(title=input_file_name_noext)
                for row in input_sheet.iter_rows():
                    for cell in row:
                        output_sheet[cell.coordinate].value = cell.value

    output_book.remove(output_book['Sheet'])
    output_book.save(os.path.join(output_folder_path, insert_datetime_before_extension(output_file_name)))

def insert_datetime_before_extension(filename):
    dt_now_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    ftitle, fext = os.path.splitext(filename)
    return ftitle + "_" + dt_now_str + fext


# テスト用例
input_folder_path = r'C:\myworkspace\python\aggregate_excel_sheets\excel'
input_sheet_name = r'Sheet1'
output_folder_path = r'C:\myworkspace\python\aggregate_excel_sheets'
output_file_name = r'output.xlsx'
aggregate_excel_sheets(input_folder_path,input_sheet_name,output_folder_path,output_file_name)
