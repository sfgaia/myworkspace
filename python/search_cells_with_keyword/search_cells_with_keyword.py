import os
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def search_cells_with_keyword(input_forlder_path, keywords,output_forlder_path,output_file_name):
    # 結果を格納するための新しいExcelブックを作成
    result_book = Workbook()
    result_sheet = result_book.active
    result_sheet.title = "result"
    result_sheet.append(["search_word:",' '.join(keywords) ])
    result_sheet.append(["file_name","sheet_name", "hit_keywords","cell_range","cell_value"])

    # 指定されたフォルダ内の.xlsxファイルを走査
    for file_name in os.listdir(input_forlder_path):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(input_forlder_path, file_name)

            # Excelファイルを開く
            workbook = load_workbook(file_path)

            # シートごとに検索
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # シート内のセルを走査し、指定されたキーワードを含むセルを探す
                for row in sheet.iter_rows():
                    for cell in row:
                        hitKeywords = []
                        for keyword in keywords:
                            if cell.value is not None and keyword in str(cell.value):
                                hitKeywords.append(keyword)
                        if hitKeywords:
                            # 結果を新しいExcelブックに書き込む
                            result_sheet.append([file_name,sheet_name, ' '.join(hitKeywords),cell.coordinate,cell.value])

            # Excelファイルを閉じる
            workbook.close()

    # 結果を保存する
    result_book.save(os.path.join(output_forlder_path, insert_datetime_before_extension(output_file_name)))
    result_book.close()

def insert_datetime_before_extension(filename):
    dt_now_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    ftitle, fext = os.path.splitext(filename)
    return ftitle + "_" + dt_now_str + fext

# テスト用例
input_forlder_path = r"C:\myworkspace\python\search_cells_with_keyword\excel"
keywords = ["test","happy"]
output_forlder_path = r"C:\myworkspace\python\search_cells_with_keyword"
output_file_name = "output.xlsx"
search_cells_with_keyword(input_forlder_path, keywords,output_forlder_path,output_file_name)
