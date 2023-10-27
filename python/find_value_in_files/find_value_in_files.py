import os
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def find_value_in_files(folder_path, sheet_name, cell_position,output_folder_path,output_file_name):
    # 新しいExcelブックを作成
    output_book = Workbook()
    output_sheet = output_book.active
    output_sheet.title = "result"

    # ヘッダを設定
    output_sheet['A1'] = "ファイル名"
    output_sheet['B1'] = "シート名"
    output_sheet['C1'] = "セル"
    output_sheet['D1'] = "値"

    # フォルダ内のファイルを走査
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(folder_path, file_name)

            # Excelファイルを読み込み
            workbook = load_workbook(file_path)

            # 指定されたシート名が存在するかチェック
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # 指定されたセル位置の値を取得
                cell_value = sheet[cell_position].value

                # 結果を出力
                output_sheet.append([file_name, sheet_name, cell_position,cell_value])
            else:
                # シート名が存在しない場合は、空白を出力
                output_sheet.append([file_name, '','', ''])

            workbook.close()

    # 結果を保存
    output_book.save(os.path.join(output_folder_path, insert_datetime_before_extension(output_file_name)))

def insert_datetime_before_extension(filename):
    dt_now_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    ftitle, fext = os.path.splitext(filename)
    return ftitle + "_" + dt_now_str + fext

# フォルダパス、シート名、セル位置を指定して関数を呼び出す
folder_path = r'C:\myworkspace\python\find_value_in_files\excel'
sheet_name = 'Sheet1'
cell_position = 'B2'
output_folder_path = r'C:\myworkspace\python\find_value_in_files'
output_file_name = r'output.xlsx'

find_value_in_files(folder_path, sheet_name, cell_position,output_folder_path,output_file_name)
