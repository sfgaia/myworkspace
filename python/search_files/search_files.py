import os
import openpyxl
import datetime

def search_files(folder_path, extensions, keyword, output_folder, output_file):
    # Excelファイルの作成
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "フォルダ名"
    sheet.cell(row=1, column=2).value = "ファイル名"
    sheet.cell(row=1, column=3).value = "行数"
    sheet.cell(row=1, column=4).value = "行の値"
    row_count = 2

    # 拡張子のリストを小文字に変換
    extensions = [ext.lower() for ext in extensions]

    # フォルダ内のファイルを再帰的に検索
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            # 拡張子が指定されたリストに含まれているか確認
            if file.lower().endswith(tuple(extensions)):
                file_path = os.path.join(root, file)
                with open(file_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                    for line_num, line in enumerate(lines, start=1):
                        # キーワードが行に含まれているか確認
                        if keyword in line:
                            sheet.cell(row=row_count, column=1).value = root
                            sheet.cell(row=row_count, column=2).value = file
                            sheet.cell(row=row_count, column=3).value = line_num
                            sheet.cell(row=row_count, column=4).value = line
                            row_count += 1

    # 結果をExcelファイルに保存
    output_path = os.path.join(output_folder, insert_datetime_before_extension(output_file))
    wb.save(output_path)
    print(f"検索結果が {output_path} に保存されました。")

def insert_datetime_before_extension(filename):
    dt_now_str = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    ftitle, fext = os.path.splitext(filename)
    return ftitle + "_" + dt_now_str + fext

# 使用例
folder_path = r"C:\myworkspace\python\search_files\test"  # 検索対象のフォルダパス
extensions = [".txt", ".csv"]  # 検索対象の拡張子のリスト
keyword = "John"  # 検索するキーワード
output_folder = r"C:\myworkspace\python\search_files"  # 出力先フォルダパス
output_file = "output.xlsx"  # 出力ファイル名

search_files(folder_path, extensions, keyword, output_folder, output_file)
